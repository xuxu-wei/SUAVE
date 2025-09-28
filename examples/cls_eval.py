# -*- coding: utf-8 -*-
"""
增强版评估与导出：
- 将单次/多次评估的结果写入**一个Excel文件**的不同sheet。
- 兼容二分类与多分类；鲁棒处理概率列缺失/不和为1等情况。
- 可配置label列、预测列、概率列前缀、阳性类等。
"""

import atexit
import glob
import inspect
import os
import re
import sys
import tempfile
from datetime import datetime
from os import PathLike
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple, Union

import numpy as np
import pandas as pd
# sklearn 依赖
from sklearn.metrics import (
    accuracy_score,
    precision_recall_fscore_support,
    balanced_accuracy_score,
    confusion_matrix,
    roc_auc_score,
    average_precision_score,
)
from sklearn.preprocessing import label_binarize
from tqdm.auto import tqdm

_PREVIEW_TMPDIR = None

DEFAULT_LABEL_COL_NAME = "label"
DEFAULT_Y_PRED_COL_NAME = "y_pred"
DEFAULT_PRED_PROBA_COL_PREFIX = "pred_proba_"

# ------------------------- 核心工具函数 -------------------------
def _extract_classes_and_probas(
    df: pd.DataFrame,
    label_col: str = DEFAULT_LABEL_COL_NAME,
    proba_prefix: str = DEFAULT_PRED_PROBA_COL_PREFIX,
) -> Tuple[List[str], Optional[pd.DataFrame], List[str]]:
    """
    自动识别类别与概率列，返回：classes, 对齐后的proba_df(按classes列顺序), warnings
    若缺失概率列，返回 proba_df=None 并记录warning。
    """
    warnings = []
    if label_col not in df.columns:
        raise ValueError(f"'{label_col}' column not found in the data.")
    label_values = df[label_col].astype(str)
    classes_from_labels = sorted(label_values.unique().tolist())

    proba_cols = [c for c in df.columns if c.startswith(proba_prefix)]
    classes_from_proba = [c.replace(proba_prefix, "", 1) for c in proba_cols]

    classes = sorted(set(classes_from_labels).union(set(classes_from_proba)))
    if not classes:
        raise ValueError("Could not infer any classes from labels or probability columns.")

    proba_df = None
    if proba_cols:
        proba_df = pd.DataFrame(index=df.index, columns=classes, dtype=float)
        for cls in classes:
            col_name = f"{proba_prefix}{cls}"
            if col_name in df.columns:
                proba_df[cls] = pd.to_numeric(df[col_name], errors="coerce")
            else:
                proba_df[cls] = np.nan
        # 二分类：若仅有一个概率列，自动补齐其互补
        if len(classes) == 2:
            n_missing_cols = proba_df.isna().all().sum()
            if n_missing_cols == 1:
                missing_cls = [c for c in classes if proba_df[c].isna().all()][0]
                present_cls = [c for c in classes if not proba_df[c].isna().all()][0]
                proba_df[missing_cls] = 1.0 - proba_df[present_cls]
                warnings.append(
                    f"Binary case with a single probability column detected. "
                    f"Filled {proba_prefix}{missing_cls} as 1 - {proba_prefix}{present_cls}."
                )
        # 行内重标化（若存在缺失/不和为1）
        proba_df_filled = proba_df.fillna(0.0)
        with np.errstate(invalid="ignore", divide="ignore"):
            row_sum = proba_df_filled.sum(axis=1)
            renorm = proba_df_filled.div(row_sum.replace(0, np.nan), axis=0)
        proba_df = renorm.fillna(0.0)
        diff = (proba_df.sum(axis=1) - 1.0).abs()
        if (diff > 1e-3).any():
            warnings.append("Some probability rows did not sum to 1 even after normalization; results may be affected.")
    else:
        warnings.append("No probability columns found. AUC and PR-AUC will not be computed.")

    return classes, proba_df, warnings


def _per_class_confusion_stats(y_true: np.ndarray, y_pred: np.ndarray, classes: List[str]) -> pd.DataFrame:
    """
    基于“一对其余”计算逐类 TP/FP/FN/TN 与 specificity。
    """
    labels = classes
    cm = confusion_matrix(y_true, y_pred, labels=labels)
    totals = cm.sum()
    per_class = []
    for idx, cls in enumerate(labels):
        TP = cm[idx, idx]
        FP = cm[:, idx].sum() - TP
        FN = cm[idx, :].sum() - TP
        TN = totals - TP - FP - FN
        specificity = TN / (TN + FP) if (TN + FP) > 0 else np.nan
        per_class.append({"class": cls, "TP": TP, "FP": FP, "FN": FN, "TN": TN, "specificity": specificity})
    return pd.DataFrame(per_class)


def _confusion_matrix_df(y_true: np.ndarray, y_pred: np.ndarray, classes: List[str]) -> pd.DataFrame:
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    df_cm = pd.DataFrame(cm, index=[f"true_{c}" for c in classes], columns=[f"pred_{c}" for c in classes])
    return df_cm

def _bootstrap_sample_indices(
    y_true: np.ndarray,
    classes: List[str],
    rng: np.random.Generator,
    strategy: str = "stratified",
) -> np.ndarray:
    """
    返回一次 bootstrap 的样本下标（长度 = len(y_true)）。
    strategy:
      - 'simple'         : 全体上有放回抽样
      - 'stratified'     : 按原始类分布在每个类内有放回抽样（保持原分布）
      - 'class_balanced' : 在“出现过样本的类”之间平均分配样本量，再在各类内有放回抽样（平衡各类权重）
                           若某类在 y_true 中没有样本，则不分配样本给该类（只在出现过的类之间平分）。
    """
    N = len(y_true)
    if N == 0:
        return np.array([], dtype=int)

    idx_all = np.arange(N)
    by_class_idx = {cls: np.where(y_true == cls)[0] for cls in classes}

    if strategy == "simple":
        return rng.choice(idx_all, size=N, replace=True)

    if strategy == "stratified":
        parts = []
        for cls in classes:
            pool = by_class_idx.get(cls, np.array([], dtype=int))
            k = len(pool)
            if k == 0:
                continue
            parts.append(rng.choice(pool, size=k, replace=True))
        return np.concatenate(parts) if parts else np.array([], dtype=int)

    if strategy == "class_balanced":
        present = [cls for cls in classes if len(by_class_idx.get(cls, [])) > 0]
        if not present:
            return rng.choice(idx_all, size=N, replace=True)  # 兜底
        Kp = len(present)
        base = N // Kp
        rem  = N - base * Kp
        counts = {cls: base + (i < rem) for i, cls in enumerate(present)}
        parts = []
        for cls in present:
            pool = by_class_idx[cls]
            parts.append(rng.choice(pool, size=counts[cls], replace=True))
        return np.concatenate(parts) if parts else np.array([], dtype=int)

    # 未知策略：回退 simple
    return rng.choice(idx_all, size=N, replace=True)

def evaluate_predictions(
    df: pd.DataFrame,
    label_col: str = "label",
    pred_col: str = "y_pred",
    positive_label: Optional[str] = None,
    proba_prefix: str = "pred_proba_",
    *,
    bootstrap_n: int = 1000,                          # 设为 0 可关闭
    ci_percentiles: Tuple[float, float] = (2.5, 97.5),# 百分位法 CI
    bootstrap_strategy: str = "stratified",  # "stratified" | "simple" | "class_balanced"
    random_state: Optional[int] = 20201021,           # 复现性
    show_progress: bool = False,                      # 是否展示 bootstrap 进度条
    progress_desc: Optional[str] = None,              # 进度条描述文本
) -> Dict[str, pd.DataFrame]:
    """
    评估分类模型预测（兼容二分类与多分类），并在存在概率列时补充 ROC-AUC / PR-AUC。

    Parameters
    ----------
    df : pandas.DataFrame
        评估数据表。至少包含真实标签列 `label_col` 与预测标签列 `pred_col`。
        若要计算 AUC/PR-AUC，应包含若干概率列，命名为
        ``{proba_prefix}{class_name}``，例如：``pred_proba_0``、``pred_proba_1``。
    label_col : str, default "label"
        真实标签列名。
    pred_col : str, default "y_pred"
        预测标签（已阈值化/Top-1）的列名。
    positive_label : str, optional
        二分类时的阳性类名称。若未提供，优先取字符串 "1"，否则取类别名排序后的最后一个。
    proba_prefix : str, default "pred_proba_"
        概率列前缀。用于自动识别 ``pred_proba_{class}`` 列。
    bootstrap_n : int, default 1000
        自助抽样次数；设为 0 可关闭 CI 计算。
    ci_percentiles : (low, high), default (2.5, 97.5)
        百分位法区间端点（百分数）。
    bootstrap_strategy : {"stratified","simple","class_balanced"}, default "stratified"
        分层（各类内重采样，保持类占比）或整体重采样。
    random_state : int | None, default 20201021
        随机种子（None 时不固定）。
    show_progress : bool, default False
        是否在 bootstrap 阶段显示 tqdm 进度条。
    progress_desc : str | None, default None
        进度条描述文本；为 None 时使用 "Bootstrap"。
    Returns
    -------
    dict of pandas.DataFrame
        包含以下键：
        - ``overall`` : 单行总体指标，含 accuracy、balanced_accuracy、macro/weighted/micro 的
          precision/recall/f1、macro specificity 等；
          二分类时还包含 ``sensitivity_pos`` / ``specificity_pos``、``roc_auc``、``pr_auc``。
          多分类时包含 ``roc_auc_macro/weighted``、``pr_auc_macro/weighted``。
        - ``per_class`` : 逐类指标，含 precision、sensitivity_recall（即 recall/TPR）、
          f1、specificity（TNR），以及每类的 ``roc_auc_ovr`` / ``pr_auc_ovr``（一对其余）。
        - ``confusion`` : 混淆矩阵（行为 ``true_{class}``、列为 ``pred_{class}``）。
        - ``class_distribution`` : 每类真实/预测样本数。
        - ``warnings`` : 计算过程中的告警信息（例如概率行和不为 1 的重标化提示）。
        - ``bootstrap_overall_records`` : 每次 bootstrap 的总体指标原始记录，含
          ``iteration`` 序号列。
        - ``bootstrap_per_class_records`` : 每次 bootstrap、每个类别的逐类指标原始记录，
          含 ``iteration`` 与 ``class`` 列。

    Raises
    ------
    ValueError
        当缺失必需列（`label_col` 或 `pred_col`），或无法从数据中识别任何类别时。
    RuntimeError
        当底层计算出现不可恢复错误时（极少见）。

    See Also
    --------
    export_to_excel : 将评估结果写入一个 Excel 的多 Sheet。
    evaluate_and_export : 批量评估文件并合并导出。

    Notes
    -----
    - **Sensitivity == Recall（TPR）**；**Specificity == TNR**（按“一对其余”逐类计算）。
    - 若仅提供二分类中的一个概率列（如正类概率），函数会补齐负类概率为 ``1 - p``。
    - 若概率行和不为 1，会在行内按已知列**重标化**，并在 ``warnings`` 中提示。
    - 若完全缺失概率列，将跳过 AUC/PR-AUC 的计算，但阈值化指标仍可计算。

    Examples
    --------
    >>> res = evaluate_predictions(df, label_col="label", pred_col="y_pred")
    >>> res["overall"].round(4)
    >>> res["per_class"].head()
    """
    # 基本检查
    if label_col not in df.columns or pred_col not in df.columns:
        missing = [c for c in [label_col, pred_col] if c not in df.columns]
        raise ValueError(f"Missing required columns: {missing}")

    y_true = df[label_col].astype(str).values
    y_pred = df[pred_col].astype(str).values

    classes, proba_df, warn_list = _extract_classes_and_probas(df, label_col=label_col, proba_prefix=proba_prefix)

    n_classes = len(classes)
    is_binary = n_classes == 2

    # 阳性类：默认优先 "1"，否则按排序后的最后一个类
    if positive_label is None:
        positive_label = "1" if "1" in classes else classes[-1]

    # -------------------- 点估计（与原版一致） --------------------
    overall: Dict[str, float] = {}
    overall["n_samples"] = len(y_true)
    overall["n_classes"] = n_classes
    overall["accuracy"] = accuracy_score(y_true, y_pred)
    overall["balanced_accuracy"] = balanced_accuracy_score(y_true, y_pred)

    prec_macro, rec_macro, f1_macro, _ = precision_recall_fscore_support(y_true, y_pred, average="macro", zero_division=0)
    prec_weighted, rec_weighted, f1_weighted, _ = precision_recall_fscore_support(y_true, y_pred, average="weighted", zero_division=0)
    prec_micro, rec_micro, f1_micro, _ = precision_recall_fscore_support(y_true, y_pred, average="micro", zero_division=0)
    overall.update({
        "precision_macro": prec_macro,
        "recall_macro": rec_macro,
        "f1_macro": f1_macro,
        "precision_weighted": prec_weighted,
        "recall_weighted": rec_weighted,
        "f1_weighted": f1_weighted,
        "precision_micro": prec_micro,
        "recall_micro": rec_micro,
        "f1_micro": f1_micro,
    })

    # 逐类 PRF + specificity
    prec_per, rec_per, f1_per, support_per = precision_recall_fscore_support(
        y_true, y_pred, labels=classes, average=None, zero_division=0
    )
    prf_df = pd.DataFrame({
        "class": classes,
        "support": support_per,
        "precision": prec_per,
        "sensitivity_recall": rec_per,  # sensitivity == recall
        "f1": f1_per,
    })
    spec_df = _per_class_confusion_stats(y_true, y_pred, classes)
    per_class_df = prf_df.merge(spec_df[["class", "specificity"]], on="class", how="left")

    # 宏平均 specificity
    overall["specificity_macro"] = float(per_class_df["specificity"].mean()) if not per_class_df.empty else np.nan

    # 二分类：记录指定阳性类的 sensitivity/specificity
    if is_binary and positive_label in classes:
        sens_pos = per_class_df.loc[per_class_df["class"] == positive_label, "sensitivity_recall"].values[0]
        spec_pos = per_class_df.loc[per_class_df["class"] == positive_label, "specificity"].values[0]
        overall["sensitivity_pos"] = sens_pos
        overall["specificity_pos"] = spec_pos
    overall["positive_label"] = positive_label

    # 概率类指标：ROC-AUC 与 PR-AUC
    y_score = None
    if proba_df is not None:
        y_score = proba_df[classes].values
        if is_binary:
            y_true_bin = (y_true == positive_label).astype(int)
            pos_idx = classes.index(positive_label)
            try:
                overall["roc_auc"] = roc_auc_score(y_true_bin, y_score[:, pos_idx])
            except Exception:
                overall["roc_auc"] = np.nan
                warn_list.append("Failed to compute ROC-AUC for binary case.")
            try:
                overall["pr_auc"] = average_precision_score(y_true_bin, y_score[:, pos_idx])
            except Exception:
                overall["pr_auc"] = np.nan
                warn_list.append("Failed to compute PR-AUC (Average Precision) for binary case.")

            # 逐类 OvR AUC/AP
            auc_per_class, ap_per_class = [], []
            for i, cls in enumerate(classes):
                y_bin = (y_true == cls).astype(int)
                try:
                    auc_val = roc_auc_score(y_bin, y_score[:, i])
                except Exception:
                    auc_val = np.nan
                try:
                    ap_val = average_precision_score(y_bin, y_score[:, i])
                except Exception:
                    ap_val = np.nan
                auc_per_class.append(auc_val)
                ap_per_class.append(ap_val)
            per_class_df["roc_auc_ovr"] = auc_per_class
            per_class_df["pr_auc_ovr"] = ap_per_class
        else:
            Y = label_binarize(y_true, classes=classes)
            try:
                overall["roc_auc_macro"] = roc_auc_score(Y, y_score, average="macro", multi_class="ovr")
                overall["roc_auc_weighted"] = roc_auc_score(Y, y_score, average="weighted", multi_class="ovr")
            except Exception:
                overall["roc_auc_macro"] = np.nan
                overall["roc_auc_weighted"] = np.nan
                warn_list.append("Failed to compute multiclass ROC-AUC.")
            try:
                overall["pr_auc_macro"] = average_precision_score(Y, y_score, average="macro")
                overall["pr_auc_weighted"] = average_precision_score(Y, y_score, average="weighted")
            except Exception:
                overall["pr_auc_macro"] = np.nan
                overall["pr_auc_weighted"] = np.nan
                warn_list.append("Failed to compute multiclass PR-AUC (Average Precision).")

            auc_per_class, ap_per_class = [], []
            for i, cls in enumerate(classes):
                y_bin = Y[:, i]
                try:
                    auc_val = roc_auc_score(y_bin, y_score[:, i])
                except Exception:
                    auc_val = np.nan
                try:
                    ap_val = average_precision_score(y_bin, y_score[:, i])
                except Exception:
                    ap_val = np.nan
                auc_per_class.append(auc_val)
                ap_per_class.append(ap_val)
            per_class_df["roc_auc_ovr"] = auc_per_class
            per_class_df["pr_auc_ovr"] = ap_per_class

    # 混淆矩阵与类分布
    confusion_df = _confusion_matrix_df(y_true, y_pred, classes)
    class_dist_df = pd.DataFrame({
        "class": classes,
        "n_true": pd.Series(y_true).value_counts().reindex(classes, fill_value=0).values,
        "n_pred": pd.Series(y_pred).value_counts().reindex(classes, fill_value=0).values,
    })

    overall_df = pd.DataFrame([overall])
    overall_bootstrap_df = pd.DataFrame()
    per_class_bootstrap_df = pd.DataFrame()

    # -------------------- Bootstrap 置信区间 --------------------
    if bootstrap_n and bootstrap_n > 0:
        rng = np.random.default_rng(random_state)

        # 需要计算 CI 的 overall 指标键
        overall_ci_keys = [
            "accuracy","balanced_accuracy",
            "precision_macro","recall_macro","f1_macro",
            "precision_weighted","recall_weighted","f1_weighted",
            "precision_micro","recall_micro","f1_micro",
            "specificity_macro",
        ]
        if is_binary:
            # pos 类 TPR/TNR
            overall_ci_keys += ["sensitivity_pos","specificity_pos"]
            if y_score is not None:
                overall_ci_keys += ["roc_auc","pr_auc"]
        else:
            if y_score is not None:
                overall_ci_keys += ["roc_auc_macro","roc_auc_weighted","pr_auc_macro","pr_auc_weighted"]

        # per-class 的指标键
        per_class_metrics = ["precision", "sensitivity_recall", "f1", "specificity"]
        if y_score is not None:
            per_class_metrics += ["roc_auc_ovr", "pr_auc_ovr"]

        overall_boot = {k: [] for k in overall_ci_keys}
        per_class_boot = {cls: {m: [] for m in per_class_metrics} for cls in classes}

        iterator = range(bootstrap_n)
        progress_bar = None
        if show_progress and bootstrap_n > 0:
            # tqdm 可视化当前模型/数据集的 bootstrap 进度。
            progress_bar = tqdm(iterator, desc=progress_desc or "Bootstrap", leave=False)
            iterator = progress_bar
        for _ in iterator:
            sample_idx = _bootstrap_sample_indices(
                y_true=y_true,
                classes=classes,
                rng=rng,
                strategy=bootstrap_strategy,
            )

            # 若样本为空（极端异常），跳过
            if sample_idx.size == 0:
                continue

            y_true_b = y_true[sample_idx]
            y_pred_b = y_pred[sample_idx]
            y_score_b = y_score[sample_idx, :] if y_score is not None else None

            # overall（阈值化）
            overall_boot["accuracy"].append(accuracy_score(y_true_b, y_pred_b))
            overall_boot["balanced_accuracy"].append(balanced_accuracy_score(y_true_b, y_pred_b))

            m_pr, m_rc, m_f1, _ = precision_recall_fscore_support(y_true_b, y_pred_b, average="macro", zero_division=0)
            overall_boot["precision_macro"].append(m_pr)
            overall_boot["recall_macro"].append(m_rc)
            overall_boot["f1_macro"].append(m_f1)

            w_pr, w_rc, w_f1, _ = precision_recall_fscore_support(y_true_b, y_pred_b, average="weighted", zero_division=0)
            overall_boot["precision_weighted"].append(w_pr)
            overall_boot["recall_weighted"].append(w_rc)
            overall_boot["f1_weighted"].append(w_f1)

            mi_pr, mi_rc, mi_f1, _ = precision_recall_fscore_support(y_true_b, y_pred_b, average="micro", zero_division=0)
            overall_boot["precision_micro"].append(mi_pr)
            overall_boot["recall_micro"].append(mi_rc)
            overall_boot["f1_micro"].append(mi_f1)

            # per-class PRF + specificity
            pr_b, rc_b, f1_b, _ = precision_recall_fscore_support(y_true_b, y_pred_b, labels=classes, average=None, zero_division=0)
            spec_b = _per_class_confusion_stats(y_true_b, y_pred_b, classes)["specificity"].values
            for i, cls in enumerate(classes):
                per_class_boot[cls]["precision"].append(pr_b[i])
                per_class_boot[cls]["sensitivity_recall"].append(rc_b[i])
                per_class_boot[cls]["f1"].append(f1_b[i])
                per_class_boot[cls]["specificity"].append(spec_b[i])

            overall_boot["specificity_macro"].append(float(np.nanmean(spec_b)))

            # AUC / AP
            if y_score_b is not None:
                if is_binary:
                    y_true_bin_b = (y_true_b == positive_label).astype(int)
                    pos_idx = classes.index(positive_label)
                    try:
                        overall_boot["roc_auc"].append(roc_auc_score(y_true_bin_b, y_score_b[:, pos_idx]))
                    except Exception:
                        overall_boot["roc_auc"].append(np.nan)
                    try:
                        overall_boot["pr_auc"].append(average_precision_score(y_true_bin_b, y_score_b[:, pos_idx]))
                    except Exception:
                        overall_boot["pr_auc"].append(np.nan)

                    for i, cls in enumerate(classes):
                        y_bin = (y_true_b == cls).astype(int)
                        try:
                            auc_val = roc_auc_score(y_bin, y_score_b[:, i])
                        except Exception:
                            auc_val = np.nan
                        try:
                            ap_val = average_precision_score(y_bin, y_score_b[:, i])
                        except Exception:
                            ap_val = np.nan
                        per_class_boot[cls]["roc_auc_ovr"].append(auc_val)
                        per_class_boot[cls]["pr_auc_ovr"].append(ap_val)

                    if positive_label in classes:
                        i_pos = classes.index(positive_label)
                        overall_boot["sensitivity_pos"].append(rc_b[i_pos])
                        overall_boot["specificity_pos"].append(spec_b[i_pos])
                else:
                    Yb = label_binarize(y_true_b, classes=classes)
                    try:
                        overall_boot["roc_auc_macro"].append(roc_auc_score(Yb, y_score_b, average="macro", multi_class="ovr"))
                    except Exception:
                        overall_boot["roc_auc_macro"].append(np.nan)
                    try:
                        overall_boot["roc_auc_weighted"].append(roc_auc_score(Yb, y_score_b, average="weighted", multi_class="ovr"))
                    except Exception:
                        overall_boot["roc_auc_weighted"].append(np.nan)
                    try:
                        overall_boot["pr_auc_macro"].append(average_precision_score(Yb, y_score_b, average="macro"))
                    except Exception:
                        overall_boot["pr_auc_macro"].append(np.nan)
                    try:
                        overall_boot["pr_auc_weighted"].append(average_precision_score(Yb, y_score_b, average="weighted"))
                    except Exception:
                        overall_boot["pr_auc_weighted"].append(np.nan)

                    for i, cls in enumerate(classes):
                        y_bin = Yb[:, i]
                        try:
                            auc_val = roc_auc_score(y_bin, y_score_b[:, i])
                        except Exception:
                            auc_val = np.nan
                        try:
                            ap_val = average_precision_score(y_bin, y_score_b[:, i])
                        except Exception:
                            ap_val = np.nan
                        per_class_boot[cls]["roc_auc_ovr"].append(auc_val)
                        per_class_boot[cls]["pr_auc_ovr"].append(ap_val)

        if progress_bar is not None:
            progress_bar.close()

        # 百分位法聚合
        lo, hi = ci_percentiles
        def _ci(x):
            x = np.asarray(x, dtype=float)
            if x.size == 0 or np.all(np.isnan(x)):
                return (np.nan, np.nan)
            return tuple(np.nanpercentile(x, [lo, hi]))

        # overall CI 列
        for k, arr in overall_boot.items():
            low, high = _ci(arr)
            overall_df[k + "_ci_low"]  = low
            overall_df[k + "_ci_high"] = high

        # per-class CI 列
        per_class_bootstrap_frames: List[pd.DataFrame] = []
        for cls in classes:
            mask = (per_class_df["class"].astype(str) == str(cls))
            for m, arr in per_class_boot[cls].items():
                low, high = _ci(arr)
                per_class_df.loc[mask, m + "_ci_low"]  = low
                per_class_df.loc[mask, m + "_ci_high"] = high
            class_boot_df = pd.DataFrame(per_class_boot[cls])
            class_boot_df.insert(0, "iteration", np.arange(len(class_boot_df)))
            class_boot_df.insert(0, "class", cls)
            per_class_bootstrap_frames.append(class_boot_df)

        overall_bootstrap_df = pd.DataFrame(overall_boot)
        if not overall_bootstrap_df.empty:
            overall_bootstrap_df.insert(
                0, "iteration", np.arange(len(overall_bootstrap_df))
            )
        per_class_bootstrap_df = (
            pd.concat(per_class_bootstrap_frames, ignore_index=True)
            if per_class_bootstrap_frames
            else pd.DataFrame(columns=["iteration", "class"])
        )

    warnings_df = pd.DataFrame({"message": warn_list}) if warn_list else pd.DataFrame(columns=["message"])

    return {
        "overall": overall_df,
        "per_class": per_class_df,
        "confusion": confusion_df,
        "class_distribution": class_dist_df,
        "warnings": warnings_df,
        "bootstrap_overall_records": overall_bootstrap_df,
        "bootstrap_per_class_records": per_class_bootstrap_df,
    }


# ------------------------- 导出封装 -------------------------
def _get_excel_writer(output_path: str):
    engines = ["openpyxl", "xlsxwriter"]
    last_err = None
    for eng in engines:
        try:
            writer = pd.ExcelWriter(output_path, engine=eng)
            return writer
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Cannot create Excel writer for {output_path}. Last error: {last_err}")

def _clean_name(name: str) -> str:
    illegal = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in illegal:
        name = name.replace(ch, '-')
    return name

def _make_unique_sheet_name(base: str, suffix: str, used: set) -> str:
    """
    Build a sheet name that fits Excel's 31-char limit and is unique.
    Strategy:
      - Clean illegal chars.
      - Reserve space for suffix.
      - If collision, inject a small token (e.g., ~2, ~3...) before suffix.
    """
    MAX_LEN = 31
    base = _clean_name(base)
    # guarantee suffix starts with separator for readability
    if not suffix.startswith("__"):
        suffix = "__" + suffix
    # primary attempt
    def build_name(token: str=""):
        allowed_base_len = MAX_LEN - len(suffix) - len(token)
        if allowed_base_len < 1:
            # pathological: fall back to suffix head
            return (suffix + token)[-MAX_LEN:]
        return base[:allowed_base_len] + token + suffix

    name = build_name()
    if name not in used:
        used.add(name)
        return name

    # resolve collisions by adding ~k token
    for k in range(2, 50):
        token = f"~{k}"
        name = build_name(token)
        if name not in used:
            used.add(name)
            return name

    # extreme fallback: hash tail
    import hashlib
    h = hashlib.md5(base.encode()).hexdigest()[:6]
    name = build_name(f"~{h}")
    # trim again if needed
    name = name[:MAX_LEN]
    if name in used:
        # last resort: truncate and add random-ish
        name = name[:-3] + "~X"
    used.add(name)
    return name

def write_results_to_excel_unique(
    results: Dict[str, Dict[str, pd.DataFrame]],
    output_path: str,
    include_warnings_sheet: bool = True,
    *,
    ci_label: Optional[str] = None,   # 表头中“95% CI”的标签；None 默认按 95%
    ci_decimals: int = 4,             # value/CI 显示小数位
) -> str:
    """
    固定导出 6 个 Sheet：
      - Summary   （精简 KPI 子集；见 DEFAULT_SUMMARY_SPEC）
      - overall   （全量总体指标）
      - Perclass  （纵表：每行 = dataset × class；首列同一 dataset 合并单元格）
      - Confusion （按数据集逐块堆叠：顶部合并一行标注 dataset → 原始矩阵 → 空行分隔）
      - ClassDist （布局同 Confusion，表为 [class, n_true, n_pred]）
      - Warnings  （两列：dataset, message）

    设计说明（稳健性）：
      - 对多类别/类不平衡，Summary 默认包含：balanced_accuracy、f1_macro、recall_macro、specificity_macro，
        以及优先 PR-AUC（pr_auc_macro/pr_auc）、其次 ROC-AUC（roc_auc_macro/roc_auc）。
        这些指标对类占比偏斜更稳健；accuracy 仅作参考（容易虚高）。
      - 二分类时，若可用，补充 sensitivity_pos / specificity_pos（针对阳性类）。
      
    汇总策略：
      - 所有 sheet 首列均为 `dataset`。
      - Perclass/Confusion/ClassDist 采用列 MultiIndex 宽表：每行一个数据集，列为 (类(或true/pred), 指标)。
      - 列集合取并集对齐，缺失以 NaN 填充。
      - Summary 与 overall 来源相同（overall 聚合），但 Summary 仅保留推荐 KPI 子集。
      
    Parameters
    ----------
    results : dict
        形如 {dataset_name: EvaluationResult} 的映射。每个 EvaluationResult 是
        evaluate_predictions 返回的字典，包含 overall、per_class、confusion、
        class_distribution、warnings。
    output_path : str
        输出的 Excel 路径（.xlsx）。
    include_warnings_sheet : bool, default True
        是否写出 `Warnings` Sheet。

    Returns
    -------
    str
        实际写入的 Excel 文件路径。
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # ====== Summary 指标选择（默认推荐）======
    # 更稳健反映多类别/类不平衡：
    # - balanced_accuracy：对各类召回平均
    # - f1_macro：每类 F1 等权，抗类占比偏斜
    # - recall_macro / specificity_macro：分别关注漏检/误报
    # 概率判别力：
    # - 优先 PR-AUC（极不平衡更有意义），其次 ROC-AUC
    # 二分类补充：
    # - sensitivity_pos / specificity_pos（指定阳性类的 TPR/TNR）
    # 仅作参考：
    # - accuracy（类不平衡时易虚高）
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    DEFAULT_SUMMARY_SPEC = [
        "n_samples", "n_classes",
        "balanced_accuracy",      # 更稳健（对各类召回率等权）
        "f1_macro",               # 抗类占比偏斜
        "recall_macro",           # = 宏平均 TPR，关注漏检 (等价于sensitivity)
        "specificity_macro",      # = 宏平均 TNR，控制误报
        "pr_auc",   # 概率指标优先 PR-AUC（极不平衡更有意义）
        "pr_auc_macro",   
        "roc_auc", # 其次 ROC-AUC（可辅佐 PR-AUC）
        "roc_auc_macro", # 其次 ROC-AUC（可辅佐 PR-AUC）
        # ("roc_auc_macro", "roc_auc"), # 回退写法, 有roc_auc_macro时仅用roc_auc_macro, 没有时用 roc_auc
        "sensitivity_pos",        # 二分类：阳性类 TPR
        "specificity_pos",        # 二分类：阳性类 TNR
        "accuracy",               # 参考/对比用，不稳健
    ]

    # ---------- 汇总辅助 ----------
    def _agg_overall(res_map: Dict[str, Dict[str, pd.DataFrame]]) -> pd.DataFrame:
        rows = []
        for ds, res in res_map.items():
            df = res.get("overall", pd.DataFrame())
            if df is None or df.empty:
                rows.append(pd.DataFrame([{}]).assign(dataset=ds))
            else:
                row = df.copy()
                row.insert(0, "dataset", ds)
                rows.append(row)
        out = pd.concat(rows, ignore_index=True, sort=False)
        cols = ["dataset"] + [c for c in out.columns if c != "dataset"]
        return out.loc[:, cols]

    def _format_float(x, decimals: int) -> str:
        # 数值转字符串；NaN 用破折号，避免 "nan"
        try:
            if pd.isna(x):
                return "—"
            return f"{float(x):.{decimals}f}"
        except Exception:
            return "—"

    def _combine_metrics_with_ci(
        df: pd.DataFrame,
        *,
        metrics: Optional[List[str]] = None,
        ci_label_text: str = "95%",
        decimals: int = 4,
        blank_if_all_missing: bool = False,
    ) -> Tuple[pd.DataFrame, Dict[str, str]]:
        """
        将 df 中的 <metric>, <metric>_ci_low, <metric>_ci_high 合并为单列：
        新列名: "<metric> (<ci_label_text> CI)"
        内容  : "<value> (<low>, <high>)"
        若对应 CI 不存在，则保留原列不变。
        返回 (新df, {原metric名: 合并后列名或原名})
        """
        out = df.copy()
        name_map: Dict[str, str] = {}
        # 自动发现 metrics
        if metrics is None:
            metrics = []
            for c in out.columns:
                # 基于后缀自动识别对应的 metric 名称
                suffix = "_ci_low"
                if c.endswith(suffix):
                    base = c[: -len(suffix)]
                    if base in out.columns and (base + "_ci_high") in out.columns:
                        metrics.append(base)
            # 去重并保序
            seen: Set[str] = set()
            metrics = [m for m in metrics if not (m in seen or seen.add(m))]

        for m in metrics:
            low_col  = f"{m}_ci_low"
            high_col = f"{m}_ci_high"
            if m in out.columns and low_col in out.columns and high_col in out.columns:
                # 记录原位置以保持列顺序
                pos = list(out.columns).index(m)
                value_s = out[m].apply(lambda v: _format_float(v, decimals))
                low_s   = out[low_col].apply(lambda v: _format_float(v, decimals))
                high_s  = out[high_col].apply(lambda v: _format_float(v, decimals))
                new_col = f"{m} ({ci_label_text} CI)"
                if blank_if_all_missing:
                    # 对于 value/low/high 全为缺失（'—'）的行，输出空字符串
                    combo = pd.Series(
                        [None] * len(out), index=out.index, dtype=object
                    )
                    for idx in out.index:
                        v_str = str(value_s.loc[idx])
                        l_str = str(low_s.loc[idx])
                        h_str = str(high_s.loc[idx])
                        if v_str == "—" and l_str == "—" and h_str == "—":
                            combo.loc[idx] = ""
                        else:
                            combo.loc[idx] = f"{v_str} ({l_str}–{h_str})"
                else:
                    combo = value_s.astype(str) + " (" + low_s.astype(str) + "–" + high_s.astype(str) + ")"
                # 删除旧列，再插入新列到原位置
                out = out.drop(columns=[m, low_col, high_col])
                out.insert(pos, new_col, combo)
                name_map[m] = new_col
            else:
                # 无CI，保持原名
                name_map[m] = m
        return out, name_map

    def _build_summary_from_overall(overall_df_full: pd.DataFrame) -> pd.DataFrame:
        """
        从 overall 全量表中筛选 Summary 精简 KPI；若存在 *_ci_low/_ci_high，则与指标并排输出。
        """
        keep_metrics = []
        cols_present = set(overall_df_full.columns)

        # 先解析 DEFAULT_SUMMARY_SPEC，得到选中的指标序列
        for item in DEFAULT_SUMMARY_SPEC:
            if isinstance(item, tuple):
                for cand in item:
                    if cand in cols_present:
                        keep_metrics.append(cand)
                        break
            else:
                if item in cols_present:
                    keep_metrics.append(item)

        # 组装列顺序：dataset + (metric, metric_ci_low, metric_ci_high)*
        cols = ["dataset"]
        for m in keep_metrics:
            cols.append(m)
            lo = m + "_ci_low"
            hi = m + "_ci_high"
            if lo in cols_present:
                cols.append(lo)
            if hi in cols_present:
                cols.append(hi)

        return overall_df_full.loc[:, [c for c in cols if c in overall_df_full.columns]]


    def _agg_perclass_long(res_map: Dict[str, Dict[str, pd.DataFrame]]) -> pd.DataFrame:
        """纵表：每行 = dataset × class；列为 per_class 的所有指标。"""
        frames = []
        for ds, res in res_map.items():
            pc = res.get("per_class", pd.DataFrame())
            if pc is None or pc.empty:
                frames.append(pd.DataFrame([{"dataset": ds}]))
                continue
            tmp = pc.copy()
            if "class" not in tmp.columns:
                first_col = tmp.columns[0]
                if first_col != "class":
                    tmp = tmp.rename(columns={first_col: "class"})
            tmp.insert(0, "dataset", ds)
            base = ["dataset", "class"]
            others = [c for c in tmp.columns if c not in base]
            tmp = tmp.loc[:, base + others]
            frames.append(tmp)
        return pd.concat(frames, ignore_index=True, sort=False)

    # ---------- 写入与样式（按引擎分支） ----------
    with _get_excel_writer(output_path) as writer:
        engine_name = getattr(writer, "engine", "").lower()
        use_xlsxwriter = (engine_name == "xlsxwriter")
        # 统一字体名（由 Excel/系统负责缺省回退）
        FONT_NAME = "Times New Roman"
        # —— 公共：工具函数（依赖于 writer 创建后的对象）—— #
        def _apply_global_text_style(sheet_name: str, df_shape: Optional[Tuple[int, int]] = None):
            """整表：字体 Times New Roman，左对齐。"""
            if use_xlsxwriter:
                ws = writer.sheets[sheet_name]
                base_fmt = writer.book.add_format({"font_name": FONT_NAME, "align": "left", "valign": "vcenter"})
                # 所有列统一设置格式；0..16383 覆盖 XFD 列上限
                ws.set_column(0, 16383, None, base_fmt)
                # 表头行也应用（行号 0）
                ws.set_row(0, None, base_fmt)
            else:
                # openpyxl：逐单元格设置字体与对齐
                from openpyxl.styles import Font, Alignment
                ws = writer.book[sheet_name]
                base_font = Font(name=FONT_NAME)
                base_align = Alignment(horizontal="left", vertical="center")
                max_r = ws.max_row
                max_c = ws.max_column
                for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
                    for cell in row:
                        cell.font = base_font
                        cell.alignment = base_align

        # 合并单元格（Perclass 首列 & Confusion/ClassDist 标题行）
        def _merge_cells(ws, r1, c1, r2, c2, value=None, *, is_title=False):
            """
            r/c: 1-based；is_title=True 时使用标题样式（#185ABD 背景、白色粗体、左对齐）。
            """
            if use_xlsxwriter:
                if is_title:
                    title_fmt = writer.book.add_format({
                        "font_name": FONT_NAME, "bold": True, "font_color": "#FFFFFF",
                        "bg_color": "#185ABD", "align": "left", "valign": "vcenter",
                    })
                    ws.merge_range(r1 - 1, c1 - 1, r2 - 1, c2 - 1, value, title_fmt)
                else:
                    base_fmt = writer.book.add_format({
                        "font_name": FONT_NAME, "align": "left", "valign": "vcenter"
                    })
                    ws.merge_range(r1 - 1, c1 - 1, r2 - 1, c2 - 1, value, base_fmt)
            else:
                from openpyxl.styles import Font, Alignment, PatternFill
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
                cell = ws.cell(row=r1, column=c1, value=value)
                if is_title:
                    cell.font = Font(name=FONT_NAME, b=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.fill = PatternFill(fill_type="solid", fgColor="185ABD")
                else:
                    cell.font = Font(name=FONT_NAME)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 基础写单元格（用于 Confusion/ClassDist 手动写入）
        def _write_cell(ws, r, c, value):
            if use_xlsxwriter:
                base_fmt = writer.book.add_format({"font_name": FONT_NAME, "align": "left", "valign": "vcenter"})
                ws.write(r - 1, c - 1, value, base_fmt)
            else:
                from openpyxl.styles import Font, Alignment
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = Font(name=FONT_NAME)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 写 Summary / overall / Perclass 
        overall_full = _agg_overall(results)
        summary_df   = _build_summary_from_overall(overall_full)
        perclass_df  = _agg_perclass_long(results)
        
        # ---- 整合CI到 overall 与 summary ----
        _ci_label_text = ci_label or "95%"  # 默认 95%，若在 evaluate_and_export 使用了非默认置信区间，可在 export_to_excel 调用时显式传入
        # (1) overall：对所有能找到 *_ci_low/_ci_high 的指标进行合并
        overall_formatted, _ = _combine_metrics_with_ci(
            overall_full,
            metrics=None,                     # None=自动发现所有具备CI的指标
            ci_label_text=_ci_label_text,
            decimals=ci_decimals,
            # overall 中若该指标与区间均缺失，则留空
            blank_if_all_missing=True,
        )
        
        # (2) summary：仅按 DEFAULT_SUMMARY_SPEC 选择的指标合并，并保持Summary列顺序
        def _build_summary_from_overall(overall_df_full: pd.DataFrame) -> pd.DataFrame:
            keep_metrics = []
            cols_present = set(overall_df_full.columns)
            for item in DEFAULT_SUMMARY_SPEC:
                if isinstance(item, tuple):
                    for cand in item:
                        if cand in cols_present:
                            keep_metrics.append(cand)
                            break
                else:
                    if item in cols_present:
                        keep_metrics.append(item)
            # 合并CI
            merged, name_map = _combine_metrics_with_ci(
                overall_df_full,
                metrics=keep_metrics,
                ci_label_text=_ci_label_text,
                decimals=ci_decimals,
                # Summary 中若该指标与区间均缺失，则留空
                blank_if_all_missing=True,
            )
            # 输出列顺序：dataset + 映射后的列
            out_cols = ["dataset"] + [name_map[m] for m in keep_metrics if name_map[m] in merged.columns]
            out_cols = [c for c in out_cols if c in merged.columns]
            return merged.loc[:, out_cols]
        
        summary_df = _build_summary_from_overall(overall_full)
        
        # Perclass：将 *_ci_low/_ci_high 合并到同名列
        perclass_df, _ = _combine_metrics_with_ci(
            perclass_df,
            metrics=None,
            ci_label_text=_ci_label_text,
            decimals=ci_decimals,
            blank_if_all_missing=True,
        )

        # 写入改为使用 overall_formatted
        summary_df.to_excel(writer, sheet_name="Summary",   index=False)
        overall_formatted.to_excel(writer, sheet_name="overall", index=False)
        perclass_df.to_excel(writer, sheet_name="Perclass", index=False)

        # 全局文本样式（字体+左对齐）
        _apply_global_text_style("Summary",   summary_df.shape)
        _apply_global_text_style("overall",   overall_formatted.shape)
        _apply_global_text_style("Perclass",  perclass_df.shape)

        # Perclass：首列相同 dataset 合并
        def _merge_perclass_dataset_col():
            if perclass_df is None or perclass_df.empty or "dataset" not in perclass_df.columns:
                return
            ws = writer.sheets["Perclass"]
            datasets = perclass_df["dataset"].tolist()
            # 识别连续区段
            spans = []
            start = 0
            for i in range(1, len(datasets) + 1):
                if i == len(datasets) or datasets[i] != datasets[i - 1]:
                    spans.append((start, i - 1))
                    start = i
            # DataFrame 写入后：第 1 行是表头 → 数据从第 2 行开始（1-based）
            for s, e in spans:
                if s == e:
                    continue
                first_row = 2 + s
                last_row  = 2 + e
                _merge_cells(ws, first_row, 1, last_row, 1, value=datasets[s], is_title=False)
        _merge_perclass_dataset_col()

        # —— 写 Confusion（分块 + 标题行 + 空行）—— #
        def _write_confusion_blocks():
            if use_xlsxwriter:
                ws = writer.book.add_worksheet("Confusion")
            else:
                ws = writer.book.create_sheet(title="Confusion")
            cur_row = 1
            for ds, res in results.items():
                cm = res.get("confusion", pd.DataFrame())
                if cm is None or cm.empty:
                    _merge_cells(ws, cur_row, 1, cur_row, 1, value=f"{ds} — Confusion: (empty)", is_title=True)
                    cur_row += 2
                    continue

                # 标题行（合并至矩阵宽度+1）
                true_labels = [str(i)[5:] if str(i).startswith("true_") else str(i) for i in cm.index]
                pred_labels = [str(j)[5:] if str(j).startswith("pred_") else str(j) for j in cm.columns]
                total_cols = 1 + len(pred_labels)
                _merge_cells(ws, cur_row, 1, cur_row, total_cols, value=f"{ds}", is_title=True)
                cur_row += 1

                # 表头
                _write_cell(ws, cur_row, 1, "")
                for c, p in enumerate(pred_labels, start=2):
                    _write_cell(ws, cur_row, c, p)
                cur_row += 1

                # 数据
                for r, t in enumerate(true_labels, start=0):
                    _write_cell(ws, cur_row, 1, t)
                    row_vals = cm.iloc[r, :].tolist()
                    for c, v in enumerate(row_vals, start=2):
                        _write_cell(ws, cur_row, c, v)
                    cur_row += 1

                # 空行分隔
                cur_row += 1

            # 整表再统一设置列宽样式（XlsxWriter）
            if use_xlsxwriter:
                ws.set_column(0, 16383, None, writer.book.add_format({"font_name": FONT_NAME, "align": "left", "valign": "vcenter"}))
            else:
                # openpyxl：已按单元格写入时设置了字体与对齐，此处无需再遍历
                pass

        _write_confusion_blocks()

        # —— 写 ClassDist（分块 + 标题行 + 空行）—— #
        def _write_classdist_blocks():
            if use_xlsxwriter:
                ws = writer.book.add_worksheet("ClassDist")
            else:
                ws = writer.book.create_sheet(title="ClassDist")
            cur_row = 1
            for ds, res in results.items():
                cd = res.get("class_distribution", pd.DataFrame())
                if cd is None or cd.empty:
                    _merge_cells(ws, cur_row, 1, cur_row, 1, value=f"{ds} — ClassDist: (empty)", is_title=True)
                    cur_row += 2
                    continue

                tmp = cd.copy()
                cols = ["class"] + [c for c in tmp.columns if c != "class"]
                tmp = tmp.loc[:, cols]

                # 标题行（合并至表宽）
                total_cols = len(tmp.columns)
                _merge_cells(ws, cur_row, 1, cur_row, total_cols, value=f"{ds}", is_title=True)
                cur_row += 1

                # 表头
                for c, name in enumerate(tmp.columns, start=1):
                    _write_cell(ws, cur_row, c, name)
                cur_row += 1

                # 数据
                for _, row in tmp.iterrows():
                    for c, name in enumerate(tmp.columns, start=1):
                        _write_cell(ws, cur_row, c, row[name])
                    cur_row += 1

                # 空行
                cur_row += 1

            if use_xlsxwriter:
                ws.set_column(0, 16383, None, writer.book.add_format({"font_name": FONT_NAME, "align": "left", "valign": "vcenter"}))
            else:
                pass

        _write_classdist_blocks()

        # —— 写 Warnings —— #
        if include_warnings_sheet:
            rows = []
            for ds, res in results.items():
                w = res.get("warnings", pd.DataFrame())
                if w is None or w.empty:
                    continue
                tmp = w.copy()
                if "message" not in tmp.columns and len(tmp.columns) >= 1:
                    tmp = tmp.rename(columns={tmp.columns[0]: "message"})
                tmp.insert(0, "dataset", ds)
                rows.append(tmp[["dataset", "message"]])
            warnings_df = pd.concat(rows, ignore_index=True, sort=False) if rows else pd.DataFrame(columns=["dataset", "message"])
            warnings_df.to_excel(writer, sheet_name="Warnings", index=False)
            _apply_global_text_style("Warnings", warnings_df.shape)

    return output_path
export_to_excel = write_results_to_excel_unique

def _get_preview_tmpdir() -> str:
    """
    单例 TemporaryDirectory：
    - 进程退出时 atexit 自动清理；
    - 在本进程生命周期内复用同一目录，避免创建过多临时目录。
    """
    global _PREVIEW_TMPDIR
    if _PREVIEW_TMPDIR is None:
        _PREVIEW_TMPDIR = tempfile.TemporaryDirectory(prefix="preview_")
        atexit.register(_PREVIEW_TMPDIR.cleanup)
    return _PREVIEW_TMPDIR.name

def preview_result_tables(
    dataset_name: str,
    tables: Dict[str, pd.DataFrame],
    *,
    include: Tuple[str, ...] = ("overall", "per_class", "warnings"),
    max_rows: int = 200,
    max_cols: int = 40,
    round_decimals: Optional[int] = 6,
    fallback_dir: Optional[str] = None,
) -> Dict[str, Dict[str, Optional[str]]]:
    """
    通用稳健的 DataFrame 预览器。

    Parameters
    ----------
    dataset_name : str
        数据集名称（用于标题与落盘文件名）。
    tables : dict[str, DataFrame]
        待展示的表；通常来自 evaluate_predictions 的返回。
    include : tuple[str]
        需要展示的键；默认只展示 overall / per_class / warnings（可改为全量）。
    max_rows, max_cols : int
        预览裁剪上限；避免巨表卡 UI。
    round_decimals : int | None
        浮点数四舍五入位数；None 表示不处理。
    fallback_dir : str | None
        当无 UI 可用时，CSV 落盘目录；默认使用临时目录。

    Returns
    -------
    dict
        {table_key: {"status": "displayed"|"saved"|"skipped"|"error", "path": <csv或None>, "msg": <错误或说明>}}
    """
    results: Dict[str, Dict[str, Optional[str]]] = {}
    # 判定可用的 UI 渠道
    def _get_ui_backend():
        try:
            from caas_jupyter_tools import display_dataframe_to_user  # type: ignore
            return ("caas", display_dataframe_to_user)
        except Exception:
            pass
        try:
            from IPython.display import display  # type: ignore
            return ("ipython", display)
        except Exception:
            pass
        return ("none", None)

    ui_name, ui_fn = _get_ui_backend()

    def _flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
        if isinstance(df.columns, pd.MultiIndex):
            df = df.copy()
            df.columns = [" / ".join(map(str, tup)) for tup in df.columns.to_flat_index()]
        return df

    def _clip_df(df: pd.DataFrame) -> pd.DataFrame:
        # 行列裁剪
        df2 = df.iloc[: max_rows, : max_cols].copy()
        # 浮点数统一保留位数，避免显示长尾噪声
        if round_decimals is not None and round_decimals >= 0:
            for c in df2.columns:
                if pd.api.types.is_float_dtype(df2[c]):
                    df2[c] = df2[c].round(round_decimals)
        return df2

    def _safe_title(key: str) -> str:
        # 统一标题命名
        name_map = {
            "overall": "Overall",
            "per_class": "Per-Class",
            "confusion": "Confusion",
            "class_distribution": "ClassDist",
            "warnings": "Warnings",
        }
        return f"{dataset_name} • {name_map.get(key, key)}"

    def _slug(s: str) -> str:
        keep = [ch if ch.isalnum() or ch in ("-", "_") else "-" for ch in s]
        return "".join(keep).strip("-_") or "table"

    # 逐表处理
    for key, df in tables.items():
        if include and key not in include:
            results[key] = {"status": "skipped", "path": None, "msg": "not in include list"}
            continue
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            results[key] = {"status": "skipped", "path": None, "msg": "empty dataframe"}
            continue
        if not isinstance(df, pd.DataFrame):
            results[key] = {"status": "error", "path": None, "msg": f"not a DataFrame: {type(df)}"}
            continue

        try:
            df_view = _clip_df(_flatten_columns(df))
            title = _safe_title(key)
            if ui_name == "caas":
                # CAAS
                ui_fn(title, df_view)  # type: ignore
                results[key] = {"status": "displayed", "path": None, "msg": f"displayed via caas ({len(df_view)}x{df_view.shape[1]})"}
            elif ui_name == "ipython":
                # Jupyter/IPython
                ui_fn(df_view)  # type: ignore
                results[key] = {"status": "displayed", "path": None, "msg": f"displayed via IPython ({len(df_view)}x{df_view.shape[1]})"}
            else:
                # 无 UI：落盘 CSV（临时或持久）
                if fallback_dir:
                    out_dir = fallback_dir  # 持久目录：调用方自管清理
                else:
                    out_dir = _get_preview_tmpdir()  # 临时目录：进程退出自动清理

                os.makedirs(out_dir, exist_ok=True)
                path = os.path.join(out_dir, f"{_slug(dataset_name)}__{_slug(key)}.csv")
                df_view.to_csv(path, index=False)

                results[key] = {
                    "status": "saved" if fallback_dir else "saved_temp",
                    "path": path,
                    "msg": f"{'persist dir' if fallback_dir else 'temp dir'}; saved CSV ({len(df_view)}x{df_view.shape[1]})",
                    "ephemeral": False if fallback_dir else True,  # 标注是否临时文件
                }
        except Exception as e:
            results[key] = {"status": "error", "path": None, "msg": f"{type(e).__name__}: {e}"}

    return results


def _auto_detect_ci_triplets(columns: Iterable[str]) -> List[Tuple[str, str, str]]:
    """Return ``(value, low, high)`` triplets detected from the column names."""
    cols = list(columns)
    col_set = set(cols)
    triplets: List[Tuple[str, str, str]] = []
    for name in cols:
        if name.endswith("_ci_low"):
            base = name[: -len("_ci_low")]
            hi = f"{base}_ci_high"
            if base in col_set and hi in col_set:
                triplets.append((base, name, hi))
    seen: Set[str] = set()
    ordered: List[Tuple[str, str, str]] = []
    for base, low, high in triplets:
        if base not in seen:
            ordered.append((base, low, high))
            seen.add(base)
    return ordered


def _make_three_line_sheet_name(
    dataset: str, metric: str, suffix: Optional[int] = None
) -> str:
    """Return an Excel-safe sheet name for three-line summary tables."""

    safe_dataset = re.sub(r"[:\\/?*\[\]]", " ", dataset).strip()
    safe_metric = re.sub(r"[:\\/?*\[\]]", " ", metric).strip()

    if suffix is not None and safe_dataset:
        safe_dataset = f"{safe_dataset} {suffix}"

    metric_part = f" ({safe_metric})" if safe_metric else ""
    if not safe_dataset and not metric_part:
        safe_name = "Summary"
    else:
        safe_name = f"{safe_dataset}{metric_part}" if safe_dataset else safe_metric

    if len(safe_name) <= 31:
        return safe_name

    if metric_part:
        max_dataset_len = max(0, 31 - len(metric_part))
        if safe_dataset and len(safe_dataset) > max_dataset_len:
            if suffix is not None:
                suffix_str = f" {suffix}" if safe_dataset else str(suffix)
                base_without_suffix = safe_dataset[: -len(suffix_str)]
                head_len = max(0, max_dataset_len - len(suffix_str))
                trimmed_head = base_without_suffix[:head_len].rstrip()
                safe_dataset = (trimmed_head + suffix_str).strip()
            else:
                safe_dataset = safe_dataset[:max_dataset_len].rstrip()
            safe_name = f"{safe_dataset}{metric_part}" if safe_dataset else metric_part

    safe_name = safe_name[:31]
    if not safe_name:
        fallback = safe_metric[:31] if safe_metric else "Summary"
        return fallback or "Summary"
    return safe_name


def _format_three_line_number(
    value: Union[str, float, int, None],
    *,
    decimals: int,
    use_decimals: bool,
    thousand_sep: bool,
    fill_value: str,
) -> str:
    if value is None:
        return fill_value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return fill_value
        return re.sub(r"(?<=\d)-(?!\s)(?=\d)", "–", stripped)

    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)

    if not np.isfinite(num):
        return fill_value

    if not use_decimals and float(num).is_integer():
        int_val = int(round(num))
        if thousand_sep and len(str(abs(int_val))) >= 5:
            return f"{int_val:,}"
        return str(int_val)

    int_part = str(int(abs(num))) if abs(num) >= 1 else "0"
    use_thousand = thousand_sep and len(int_part) >= 5
    fmt = f"{{:,.{decimals}f}}" if use_thousand else f"{{:.{decimals}f}}"
    return fmt.format(num)


def _format_three_line_ci(
    value: Union[str, float, int, None],
    low: Union[str, float, int, None],
    high: Union[str, float, int, None],
    *,
    decimals: int,
    thousand_sep: bool,
    fill_value: str,
) -> str:
    value_txt = _format_three_line_number(
        value,
        decimals=decimals,
        use_decimals=True,
        thousand_sep=thousand_sep,
        fill_value=fill_value,
    )
    low_txt = _format_three_line_number(
        low,
        decimals=decimals,
        use_decimals=True,
        thousand_sep=thousand_sep,
        fill_value=fill_value,
    )
    high_txt = _format_three_line_number(
        high,
        decimals=decimals,
        use_decimals=True,
        thousand_sep=thousand_sep,
        fill_value=fill_value,
    )

    if any(part == fill_value for part in (value_txt, low_txt, high_txt)):
        return fill_value
    return f"{value_txt} ({low_txt}–{high_txt})"


def _prepare_three_line_sheet(
    df: pd.DataFrame,
    *,
    index_columns: Sequence[str],
    dataset_column: str,
    dataset_order: Optional[Sequence[str]] = None,
    drop_columns: Optional[Sequence[str]] = None,
    decimals: int = 3,
    thousand_sep: bool = True,
    fill_value: str = "NA",
    ci_pairs: Optional[Sequence[Tuple[str, str, str]]] = None,
    ci_label_text: str = "95%",
    decimal_metric_overrides: Optional[Sequence[str]] = None,
) -> pd.DataFrame:
    if not index_columns:
        raise ValueError("'index_columns' must contain at least one column name")
    if dataset_column not in df.columns:
        raise ValueError(f"'{dataset_column}' column not found in dataframe")

    work = df.copy()
    if drop_columns:
        drop_list = [c for c in drop_columns if c in work.columns]
        if drop_list:
            work = work.drop(columns=drop_list)

    triplets = list(ci_pairs) if ci_pairs is not None else _auto_detect_ci_triplets(work.columns)
    triplets = [t for t in triplets if len(t) == 3]
    for base, low, high in triplets:
        if base not in work.columns or low not in work.columns or high not in work.columns:
            continue
        combined = work.apply(
            lambda row: _format_three_line_ci(
                row.get(base),
                row.get(low),
                row.get(high),
                decimals=decimals,
                thousand_sep=thousand_sep,
                fill_value=fill_value,
            ),
            axis=1,
        )
        label = f"{base} ({ci_label_text} CI)"
        if label in work.columns:
            work = work.drop(columns=[label])
        work = work.drop(columns=[low, high])
        base_pos = list(work.columns).index(base)
        work.insert(base_pos + 1, label, combined)

    if dataset_order:
        categories = list(dataset_order)
        extras = [c for c in work[dataset_column].unique() if c not in categories]
        categories.extend(sorted(extras))
        work[dataset_column] = pd.Categorical(work[dataset_column], categories=categories, ordered=True)

    sort_cols = list(index_columns) + [dataset_column]
    work = work.sort_values(sort_cols)

    value_columns = [c for c in work.columns if c not in (*index_columns, dataset_column)]
    if not value_columns:
        raise ValueError("No value columns available for three-line export")

    long_df = work.loc[:, list(index_columns) + [dataset_column] + value_columns]

    decimal_metrics: Set[str] = set(decimal_metric_overrides or [])
    for col in value_columns:
        series = pd.to_numeric(long_df[col], errors="coerce")
        if col in decimal_metrics:
            continue
        if series.notna().any():
            rounded = np.round(series.dropna().astype(float))
            if not np.allclose(series.dropna().astype(float), rounded):
                decimal_metrics.add(col)
            elif pd.api.types.is_float_dtype(long_df[col]):
                decimal_metrics.add(col)

    wide = (
        long_df.set_index(list(index_columns) + [dataset_column])[value_columns]
        .unstack(dataset_column)
        .swaplevel(axis=1)
    )
    wide = wide.reset_index()

    tuples: List[Tuple[str, str]] = []
    for col in wide.columns:
        if isinstance(col, tuple):
            tuples.append((str(col[0]), str(col[1])))
        else:
            tuples.append(("", str(col)))
    wide.columns = pd.MultiIndex.from_tuples(tuples)

    prefix_cols = [col for col in wide.columns if col[0] == ""]
    dataset_sequence: List[str] = []
    for ds in wide.columns.get_level_values(0):
        if ds and ds not in dataset_sequence:
            dataset_sequence.append(ds)
    if dataset_order:
        ordered_ds = [ds for ds in dataset_order if ds in dataset_sequence]
        ordered_ds.extend([ds for ds in dataset_sequence if ds not in ordered_ds])
    else:
        ordered_ds = dataset_sequence
    column_order: List[Tuple[str, str]] = list(prefix_cols)
    for ds in ordered_ds:
        column_order.extend([col for col in wide.columns if col[0] == ds])
    wide = wide.loc[:, column_order]

    formatted = wide.copy()
    for col in formatted.columns:
        if col[0] == "":
            formatted[col] = formatted[col].apply(lambda v: fill_value if pd.isna(v) else str(v))
            continue
        metric = col[1]
        use_decimals = metric in decimal_metrics
        formatted[col] = formatted[col].apply(
            lambda v: _format_three_line_number(
                v,
                decimals=decimals,
                use_decimals=use_decimals,
                thousand_sep=thousand_sep,
                fill_value=fill_value,
            )
        )

    return formatted.fillna(fill_value)


def _write_three_line_workbook(
    sheets: Dict[str, pd.DataFrame],
    output_path: Union[str, Path],
    *,
    font_name: str = "Times New Roman",
) -> Path:
    if not sheets:
        raise ValueError("No sheets provided for three-line export")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    medium_side = Side(style="medium", color="000000")
    thin_side = Side(style="thin", color="000000")

    def _update_border(cell, *, top=None, bottom=None):
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=top if top is not None else cell.border.top,
            bottom=bottom if bottom is not None else cell.border.bottom,
        )

    def _write_sheet(ws, name: str, data: pd.DataFrame):
        ws.title = name
        header_font = Font(name=font_name, bold=True)
        body_font = Font(name=font_name, bold=False)
        align = Alignment(horizontal="left", vertical="center")

        level0 = data.columns.get_level_values(0)
        level1 = data.columns.get_level_values(1)
        n_cols = data.shape[1]

        start_col = 1
        current = level0[0] if n_cols else ""
        for idx in range(1, n_cols + 1):
            value = level0[idx - 1]
            cell = ws.cell(row=1, column=idx, value=value if value else "")
            cell.font = header_font
            cell.alignment = align
            if value != current:
                if current and idx - start_col > 1:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=idx - 1)
                start_col = idx
                current = value
        if current and n_cols - start_col >= 1:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=n_cols)

        for idx in range(1, n_cols + 1):
            cell = ws.cell(row=2, column=idx, value=level1[idx - 1])
            cell.font = header_font
            cell.alignment = align

        for row_idx, row in enumerate(data.itertuples(index=False, name=None), start=3):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = align

        for col_idx in range(1, n_cols + 1):
            top_cell = ws.cell(row=1, column=col_idx)
            _update_border(top_cell, top=medium_side)
            header_cell = ws.cell(row=2, column=col_idx)
            _update_border(header_cell, bottom=thin_side)
            bottom_cell = ws.cell(row=data.shape[0] + 2, column=col_idx)
            _update_border(bottom_cell, bottom=medium_side)

        for col_idx in range(1, n_cols + 1):
            values = [
                str(level0[col_idx - 1] or ""),
                str(level1[col_idx - 1] or ""),
            ] + [str(v) for v in data.iloc[:, col_idx - 1].tolist()]
            max_len = max((len(v) for v in values if v is not None), default=0)
            width = min(max(max_len + 2, 10), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    first = True
    for sheet_name, table in sheets.items():
        if table.empty:
            continue
        if first:
            worksheet = workbook.active
            _write_sheet(worksheet, sheet_name, table)
            first = False
        else:
            worksheet = workbook.create_sheet(title=sheet_name)
            _write_sheet(worksheet, sheet_name, table)

    workbook.save(output_path)
    return Path(output_path)


def export_three_line_tables(
    tables: Dict[str, pd.DataFrame],
    output_path: Union[str, Path],
    *,
    index_columns: Dict[str, Sequence[str]],
    dataset_column: str = "Dataset",
    dataset_order: Optional[Sequence[str]] = None,
    drop_columns: Optional[Sequence[str]] = ("Target",),
    decimals: int = 3,
    thousand_sep: bool = True,
    fill_value: str = "NA",
    ci_pairs: Optional[Dict[str, Sequence[Tuple[str, str, str]]]] = None,
    ci_label_text: str = "95%",
    decimal_metric_overrides: Optional[Dict[str, Sequence[str]]] = None,
) -> Path:
    """Export multiple dataframes as academic three-line tables grouped by dataset.

    Examples
    --------
    >>> tables = {"Summary": summary_df, "overall": overall_df}
    >>> export_three_line_tables(
    ...     tables,
    ...     "benchmark_three_line.xlsx",
    ...     index_columns={"Summary": ["Model"], "overall": ["Model"]},
    ... )
    PosixPath('benchmark_three_line.xlsx')
    """

    prepared: Dict[str, pd.DataFrame] = {}
    for sheet_name, df in tables.items():
        if df is None or df.empty:
            continue
        idx_cols = index_columns.get(sheet_name)
        if not idx_cols:
            raise ValueError(f"Missing index columns configuration for sheet '{sheet_name}'")
        ci_spec = ci_pairs.get(sheet_name) if ci_pairs else None
        decimal_override = (
            decimal_metric_overrides.get(sheet_name)
            if decimal_metric_overrides and sheet_name in decimal_metric_overrides
            else None
        )
        prepared[sheet_name] = _prepare_three_line_sheet(
            df,
            index_columns=idx_cols,
            dataset_column=dataset_column,
            dataset_order=dataset_order,
            drop_columns=drop_columns,
            decimals=decimals,
            thousand_sep=thousand_sep,
            fill_value=fill_value,
            ci_pairs=ci_spec,
            ci_label_text=ci_label_text,
            decimal_metric_overrides=decimal_override,
        )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return _write_three_line_workbook(prepared, output_path)


def _normalize_filepaths(
    filepaths: Union[
        str, PathLike,
        List[Union[str, PathLike]],
        Tuple[Union[str, PathLike], ...],
        Set[Union[str, PathLike]],
        np.ndarray, pd.Series
    ]
) -> List[str]:
    """
    规范化输入的文件路径，返回**有序去重**后的路径列表；支持 PathLike 与通配展开。

    Parameters
    ----------
    filepaths : str, os.PathLike, list-like of (str | os.PathLike), numpy.ndarray, pandas.Series
        - 单个字符串或 Path/PathLike：若包含 ``*?[]``，将用 :mod:`glob` 展开；
          若无匹配，保留原字符串作为单文件路径。
        - 可迭代对象（list/tuple/set/ndarray/Series）中每个元素均按上述规则处理。
          不接受非字符串/PathLike 的元素（会抛出 TypeError）。

    Returns
    -------
    list of str
        规范化后的文件路径列表；**去重但保持首次出现顺序**。

    Raises
    ------
    TypeError
        当 ``filepaths`` 或其中元素不是 ``str`` / ``os.PathLike`` 时。
    ValueError
        展开与过滤后为空列表时。

    See Also
    --------
    evaluate_and_export : 读取文件、评估并导出；内部调用本函数做路径预处理。

    Notes
    -----
    - 使用 ``os.fspath`` 将 ``PathLike`` 转换为字符串，保证后续 I/O 与 glob 行为一致。
    - 通配展开使用 :func:`glob.glob`；无匹配时按原字符串保留（用于显式不存在文件的告警记录）。

    Examples
    --------
    >>> _normalize_filepaths("/mnt/data/*.xlsx")
    ['/mnt/data/A.xlsx', '/mnt/data/B.xlsx']

    >>> from pathlib import Path
    >>> _normalize_filepaths([Path("/mnt/data/A.xlsx"), "/mnt/data/B.xlsx"])
    ['/mnt/data/A.xlsx', '/mnt/data/B.xlsx']
    """

    def to_str(p) -> str:
        # 支持 pathlib.Path / 任何 os.PathLike；其余按 str() 兜底
        if isinstance(p, os.PathLike):
            return os.fspath(p)
        return str(p)

    def expand_one(p) -> List[str]:
        s = to_str(p)
        # 若包含通配符则尝试 glob 展开；无匹配则保留原样
        if any(ch in s for ch in "*?[]"):
            matches = glob.glob(s)
            return matches if matches else [s]
        return [s]

    out: List[str] = []

    if isinstance(filepaths, (str, PathLike)):
        out.extend(expand_one(filepaths))

    elif isinstance(filepaths, (list, tuple, set)):
        for fp in filepaths:
            if not isinstance(fp, (str, PathLike)):
                raise TypeError(f"All elements in filepaths must be str or os.PathLike, got {type(fp)}")
            out.extend(expand_one(fp))

    elif isinstance(filepaths, (np.ndarray, pd.Series)):
        # 逐元素转换为 PathLike/str 再展开
        seq = list(filepaths)
        for fp in seq:
            if not isinstance(fp, (str, PathLike)):
                # 容错：转字符串（例如 dtype=object 杂糅时）
                fp = to_str(fp)
            out.extend(expand_one(fp))

    else:
        raise TypeError(f"'filepaths' must be a str, os.PathLike, or a list-like of those, got {type(filepaths)}")

    # 去重但保序
    seen = set()
    deduped: List[str] = []
    for p in out:
        if p not in seen:
            seen.add(p)
            deduped.append(p)

    if not deduped:
        raise ValueError("No valid file paths provided after normalization.")

    return deduped

def _get_entry_script_dir() -> Optional[Path]:
    """
    尝试获取“入口脚本”所在目录：
    1) 首选调用栈中 __main__ frame 的 __file__
    2) 退而求其次使用 sys.argv[0]
    找不到或无文件时返回 None
    """
    # 1) 调用栈中寻找 __main__ 帧（最稳妥的“谁在跑我”）
    for frame in inspect.stack():
        g = frame.frame.f_globals
        if g.get("__name__") == "__main__":
            f = g.get("__file__")
            if f:
                p = Path(f).resolve()
                if p.exists():
                    return p.parent
    # 2) sys.argv[0]（某些 runner 会设置可用的脚本路径）
    argv0 = Path(sys.argv[0]).resolve() if sys.argv and sys.argv[0] else None
    if argv0 and argv0.exists():
        return argv0.parent
    return None


def _default_output_path(
    prefix: str = "model_eval",
    ext: str = ".xlsx",
    prefer_script_dir: bool = True,
    env_var: str = "MODEL_EVAL_OUTPUT_DIR",
) -> Path:
    """
    生成安全的默认输出路径：优先入口脚本目录 -> CWD -> 用户家目录 -> 临时目录
    支持环境变量覆盖：MODEL_EVAL_OUTPUT_DIR

    返回值为 Path；调用处可 str(...) 转换。
    """
    # 0) 环境变量覆盖（若你希望禁用该能力，可移除此块）
    env_dir = os.environ.get(env_var)
    if env_dir:
        base = Path(env_dir).expanduser().resolve()
    else:
        base = None

    # 1) 入口脚本目录（可选）
    if base is None and prefer_script_dir:
        base = _get_entry_script_dir()

    # 2) 当前工作目录
    if base is None:
        base = Path.cwd()

    # 3) 可写性检测；不行则回退用户家目录，再不行回退临时目录
    def _is_writable(p: Path) -> bool:
        try:
            p.mkdir(parents=True, exist_ok=True)
            test = p / ".write_test.tmp"
            with open(test, "w") as f:
                f.write("ok")
            test.unlink(missing_ok=True)
            return True
        except Exception:
            return False

    if not _is_writable(base):
        home = Path.home()
        base = home if _is_writable(home) else Path(tempfile.gettempdir())

    # 4) 生成文件名：时间戳避免覆盖；若极端情况下重复，追加计数
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = base / f"{prefix}_{ts}{ext}"
    if not candidate.exists():
        return candidate

    # 极短时间内多次调用的并发情形：追加序号
    for k in range(2, 1000):
        c = base / f"{prefix}_{ts}~{k}{ext}"
        if not c.exists():
            return c
    # 理论不会到这一步
    return base / f"{prefix}_{ts}~{os.getpid()}{ext}"

def evaluate_and_export(
    filepaths: Union[str, List[str], tuple, set, np.ndarray, pd.Series],
    output_path: Optional[str],
    label_col: str = DEFAULT_LABEL_COL_NAME,
    pred_col: str = DEFAULT_Y_PRED_COL_NAME,
    positive_label: Optional[str] = None,
    proba_prefix: str = DEFAULT_PRED_PROBA_COL_PREFIX,
    show_preview_tables: bool = False,
    *,
    bootstrap_n: int = 1000,
    ci_percentiles: Tuple[float, float] = (2.5, 97.5),
    bootstrap_strategy: str = "stratified",
    random_state: Optional[int] = 42,
) -> str:
    """
    读取一个或多个评估文件，执行 :func:`evaluate_predictions`，并将所有结果合并导出到**一个** Excel。

    Parameters
    ----------
    filepaths : str, list-like
        支持以下形式：
        - 单个路径字符串或 PathLike；
        - 多个路径的 list/tuple/set/ndarray/Series；
        - 含通配符的字符串（例如 ``'/mnt/data/*_eval.xlsx'``），内部使用 :func:`glob.glob` 展开；
        - 顺序将被保留，重复路径会被去重（保留首次出现）。
    output_path : str or None
        输出 Excel 的路径（建议以 ``.xlsx`` 结尾）。若为 ``None``，将自动生成形如
        ``/mnt/data/model_eval_{YYYYmmdd_HHMMSS}.xlsx`` 的路径。
    label_col : str, default "label"
        真实标签列名。
    pred_col : str, default "y_pred"
        预测标签列名。
    positive_label : str, optional
        二分类阳性类；未提供时优先取 "1"，否则取类别排序末位。
    proba_prefix : str, default "pred_proba_"
        概率列前缀；用于识别 ``{proba_prefix}{class}`` 格式的概率列。
    show_preview_tables : bool, default False
        是否在交互环境下预览每个数据集的 ``overall``/``per_class``/``warnings`` 表（若运行环境支持）。
    bootstrap_n : int, default 1000
        自助抽样次数；设为 0 可关闭 CI 计算。
    ci_percentiles : (low, high), default (2.5, 97.5)
        百分位法区间端点（百分数）。
    bootstrap_strategy : {"stratified","simple","class_balanced"}, default "stratified"
        分层（各类内重采样，保持类占比）或整体重采样。
    random_state : int | None, default 20201021
        随机种子（None 时不固定）。
        
    Returns
    -------
    str
        实际写入的 Excel 文件路径。

    Raises
    ------
    TypeError
        当 ``filepaths`` 的类型不被支持，或内部元素非 ``str/os.PathLike``。
    ValueError
        当规范化后没有任何可用路径。
    RuntimeError, OSError
        当写入 Excel 失败时（写权限/磁盘问题/Excel 引擎问题）。

    See Also
    --------
    evaluate_predictions : 单个 DataFrame 的评估。
    export_to_excel : 将已有的结果字典写入 Excel。

    Notes
    -----
    - 对每个输入文件都会生成 5 个基础 Sheet：``Overall``、``PerClass``、``Confusion``、
      ``ClassDist``、``Warnings``；另含总览 ``Summary``。
    - 支持 ``.xlsx/.xls/.csv/.tsv``；未知扩展名会被记录为 ``Warnings``，但不会中断其他文件导出。
    - Sheet 名称自动规避 Excel 限制（非法字符、31 字符上限、重名冲突）。

    Examples
    --------
    >>> evaluate_and_export(
    ...     filepaths=["/mnt/data/A.xlsx", "/mnt/data/B.xlsx"],
    ...     output_path="/mnt/data/combined_metrics.xlsx",
    ...     proba_prefix="pred_proba_",
    ... )
    '/mnt/data/combined_metrics.xlsx'
    """
    # —— 恢复你原先的“当 output_path 为空则自动命名”的逻辑 ——
    if output_path is None:
        raise ValueError("Must provide `output_path`")
        # ts = time.strftime("%Y%m%d_%H%M%S")
        # output_path = f"./model_eval_{ts}.xlsx"
        
        # output_path = str(_default_output_path(
        #     prefix="model_eval",
        #     ext=".xlsx",
        #     prefer_script_dir=True,              # 先尝试入口脚本目录
        #     env_var="MODEL_EVAL_OUTPUT_DIR",     # 可用环境变量强制覆盖
        # ))
    
    file_list = _normalize_filepaths(filepaths)

    results_all: Dict[str, Dict[str, pd.DataFrame]] = {}
    for fp in file_list:
        name = os.path.splitext(os.path.basename(fp))[0]
        if not os.path.exists(fp):
            results_all[name] = {
                "overall": pd.DataFrame(),
                "per_class": pd.DataFrame(),
                "confusion": pd.DataFrame(),
                "class_distribution": pd.DataFrame(),
                "warnings": pd.DataFrame({"message": [f"File not found: {fp}"]}),
            }
            continue

        ext = os.path.splitext(fp)[1].lower()
        if ext in [".xlsx", ".xls"]:
            df = pd.read_excel(fp)
        elif ext in [".csv", ".tsv"]:
            sep = "," if ext == ".csv" else "\t"
            df = pd.read_csv(fp, sep=sep)
        else:
            results_all[name] = {
                "overall": pd.DataFrame(),
                "per_class": pd.DataFrame(),
                "confusion": pd.DataFrame(),
                "class_distribution": pd.DataFrame(),
                "warnings": pd.DataFrame({"message": [f"Unsupported file extension: {ext}"]}),
            }
            continue

        try:
            res = evaluate_predictions(
                df=df,
                label_col=label_col,
                pred_col=pred_col,
                positive_label=positive_label,
                proba_prefix=proba_prefix,
                bootstrap_n=bootstrap_n,
                ci_percentiles=ci_percentiles,
                bootstrap_strategy=bootstrap_strategy,
                random_state=random_state,
            )
        except Exception as e:
            res = {
                "overall": pd.DataFrame(),
                "per_class": pd.DataFrame(),
                "confusion": pd.DataFrame(),
                "class_distribution": pd.DataFrame(),
                "warnings": pd.DataFrame({"message": [f"Evaluation failed for {name}: {e}"]}),
            }
        results_all[name] = res

        if show_preview_tables:
            _preview_info = preview_result_tables(
                dataset_name=name,
                tables=res,
                include=("overall", "per_class", "confusion"),
                max_rows=200,
                max_cols=40,
                round_decimals=4,
                fallback_dir=None,
            )
            # print(_preview_info)
            
    # 在 evaluate_and_export 末尾写入 Excel 前，推导CI标签并透传
    _ci_level = None
    try:
        lo, hi = ci_percentiles
        lvl = float(hi) - float(lo)
        if 0 < lvl < 100:
            _ci_level = f"{int(round(lvl))}%"
    except Exception:
        pass
    out = export_to_excel(
        results_all,
        output_path=output_path,
        include_warnings_sheet=True,
        ci_label=_ci_level or "95%",
    )
    return out

def evaluate_and_export_autoname(
    filepaths: Union[str, List[str], tuple, set, np.ndarray, pd.Series],
    output_path: Optional[str] = None,
    *args, **kwargs,
) -> str:
    """
    便捷包装：当未提供 output_path 时，自动生成一个安全的默认文件名。
    其它参数与 evaluate_and_export 一致。
    """
    if output_path is None:
        output_path = str(_default_output_path(
            prefix="model_eval", ext=".xlsx", prefer_script_dir=True, env_var="MODEL_EVAL_OUTPUT_DIR"
        ))
    return evaluate_and_export(filepaths=filepaths, output_path=output_path, *args, **kwargs)

__all__ = [
    "DEFAULT_LABEL_COL_NAME",
    "DEFAULT_Y_PRED_COL_NAME",
    "DEFAULT_PRED_PROBA_COL_PREFIX",
    "evaluate_predictions",
    "evaluate_and_export",
    "evaluate_and_export_autoname",   # 新增导出
    "export_to_excel",
    "export_three_line_tables",
]
